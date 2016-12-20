
import os.path as op

try:
    from typing import List, Sequence
except:
    raise ImportError('`typing` module not found, please install it.')

import pandas as pd
import xlrd
from openpyxl import load_workbook


def _openpyxl_read_xl(xl_path: str):
    """ Use openpyxl to read an Excel file. """
    try:
        wb = load_workbook(filename=xl_path, read_only=True)
    except:
        raise
    else:
        return wb


def _xlrd_read_xl(xl_path: str):
    """ Use xlrd to get the list of sheet names from `xl_path`."""
    try:
        wb = xlrd.open_workbook(xl_path)
    except:
        raise
    else:
        return wb


XL_READERS = {'xlrd': _xlrd_read_xl,
              'openpyxl': _openpyxl_read_xl,
             }


def _use_openpyxl_or_xlrf(xl_path: str):
    fails   = []
    choices = XL_READERS.keys()
    for m in choices:
        reader = XL_READERS[m]

        try:
            reader(xl_path)
        except:
            fails.append(m)
            pass
        else:
            return m

    raise RuntimeError("Could not open {} with {}.".format(xl_path,
                                                           ' nor '.join(fails)))


def _check_xl_path(xl_path: str):
    """ Return the expanded absolute path of `xl_path` if
    if exists and 'xlrd' or 'openpyxl' depending on
    which module should be used for the Excel file in `xl_path`.

    Parameters
    ----------
    xl_path: str
        Path to an Excel file

    Returns
    -------
    xl_path: str
        User expanded and absolute path to `xl_path`

    module: str
        The name of the module you should use to process the
        Excel file.
        Choices: 'xlrd', 'pyopenxl'

    Raises
    ------
    IOError
        If the file does not exist

    RuntimError
        If a suitable reader for xl_path is not found
    """
    xl_path = op.abspath(op.expanduser(xl_path))

    if not op.isfile(xl_path):
        raise IOError("Could not find file in {}.".format(xl_path))

    return xl_path, _use_openpyxl_or_xlrf(xl_path)


def read_xl(xl_path: str):
    """ Return the workbook from the Excel file in `xl_path`."""
    xl_path, choice = _check_xl_path(xl_path)
    reader = XL_READERS[choice]

    return reader(xl_path)


def get_sheet_list(xl_path: str) -> List:
    """Return a list with the name of the sheets in
    the Excel file in `xl_path`.
    """
    wb = read_xl(xl_path)

    if hasattr(wb, 'sheetnames'):
        return wb.sheetnames
    else:
        return wb.sheet_names()


def concat_sheets(xl_path: str, sheetnames=None, add_tab_names=False):
    """ Return a pandas DataFrame with the concat'ed
    content of the `sheetnames` from the Excel file in
    `xl_path`.

    Parameters
    ----------
    xl_path: str
        Path to the Excel file

    sheetnames: list of str
        List of existing sheet names of `xl_path`.
        If None, will use all sheets from `xl_path`.

    add_tab_names: bool
        If True will add a 'Tab' column which says from which
        tab the row comes from.

    Returns
    -------
    df: pandas.DataFrame
    """
    xl_path, choice = _check_xl_path(xl_path)

    if sheetnames is None:
        sheetnames = get_sheet_list(xl_path)

    sheets = pd.read_excel(xl_path, sheetname=sheetnames)

    if add_tab_names:
        for tab in sheets:
            sheets[tab]['Tab'] = [tab] * len(sheets[tab])

    return pd.concat([sheets[tab] for tab in sheets])


def _check_cols(df, col_names):
    """ Raise an AttributeError if `df` does not have a column named as an item of
    the list of strings `col_names`.
    """
    for col in col_names:
        if not hasattr(df, col):
            raise AttributeError("DataFrame does not have a '{}' column, got {}.".format(col,
                                                                                         df.columns))


def col_values(df, col_name):
    """ Return a list of not null values from the `col_name` column of `df`."""
    _check_cols(df, [col_name])

    if 'O' in df[col_name] or pd.np.issubdtype(df[col_name].dtype, str): # if the column is of strings
        return [nom.lower() for nom in df[pd.notnull(df)][col_name] if not pd.isnull(nom)]
    else:
        return [nom for nom in df[pd.notnull(df)][col_name] if not pd.isnull(nom)]


def duplicated_rows(df, col_name):
    """ Return a DataFrame with the duplicated values of the column `col_name`
    in `df`."""
    _check_cols(df, [col_name])

    dups = df[pd.notnull(df[col_name]) & df.duplicated(subset=[col_name])]
    return dups


def duplicated(values: Sequence):
    """ Return the duplicated items in `values`"""
    vals = pd.Series(values)
    return vals[vals.duplicated()]
